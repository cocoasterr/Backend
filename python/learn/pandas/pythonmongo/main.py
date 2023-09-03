import time
from openpyxl import load_workbook
from pymongo import MongoClient

def process_sheet(sheet):
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data

def read_excel(filename):
    wb = load_workbook(filename, data_only=True)
    all_data = []
    for sheet in wb:
        data = process_sheet(sheet)
        all_data.extend(data)
    return all_data

def insert_to_mongodb(data):
    client = MongoClient("mongodb://localhost:27017/")
    db = client["mydb"]
    collection = db["mycollection"]

    start_time = time.time()

    for row in data:
        document = {}
        for i, value in enumerate(row):
            column_name = f"column{i + 1}"
            document[column_name] = value

        collection.insert_one(document)

    end_time = time.time()
    elapsed_time = end_time - start_time
    client.close()

    return elapsed_time

if __name__ == "__main__":
    # Nama file Excel yang akan dibaca
    filename = "../Orders-With Nulls.xlsx"

    # Memulai mengukur waktu eksekusi
    total_start_time = time.time()

    # Baca Excel
    excel_data = read_excel(filename)

    # Menghentikan mengukur waktu eksekusi membaca Excel
    read_end_time = time.time()
    read_elapsed_time = read_end_time - total_start_time
    print(f"Read Executed in: {read_elapsed_time} seconds")

    # Insert ke MongoDB
    insert_duration = insert_to_mongodb(excel_data)

    # Menghentikan mengukur waktu eksekusi penyisipan
    total_end_time = time.time()
    total_elapsed_time = total_end_time - total_start_time

    print(f"Migration to Mongodb in: {insert_duration} seconds")
    print(f"Total Execute time: {total_elapsed_time} seconds")
